"""Utilities for post-processing generated Excel workbooks."""
from __future__ import annotations

import json

from openpyxl import load_workbook
from openpyxl.styles import Alignment


def post_process_workbook(workbook):
    """Apply formatting, alignment, and JSON pretty-printing to the workbook."""
    for sheet_name in workbook.sheetnames:
        worksheet = workbook[sheet_name]
        min_row_height = 18
        header_height = 24

        for col_idx, cell in enumerate(worksheet[1], 1):
            if cell.value and str(cell.value).strip().lower() == "publicip":
                worksheet.delete_cols(col_idx)
                break

        for _, cell in enumerate(worksheet[1], 1):
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        worksheet.row_dimensions[1].height = header_height

        for idx, row in enumerate(worksheet.iter_rows(min_row=2, max_row=worksheet.max_row), start=2):
            max_lines = 1
            for cell in row:
                value = cell.value
                if value and isinstance(value, str):
                    lines = value.count("\n") + 1
                    if lines > max_lines:
                        max_lines = lines
                cell.alignment = Alignment(wrap_text=True, horizontal="center", vertical="center")
            worksheet.row_dimensions[idx].height = max(15 * max_lines, min_row_height)

        for col_idx, _ in enumerate(worksheet[1], 1):
            for row in range(2, worksheet.max_row + 1):
                value = worksheet.cell(row=row, column=col_idx).value
                if value and isinstance(value, str) and (value.strip().startswith("{") or value.strip().startswith("[")):
                    try:
                        parsed = json.loads(value)
                        formatted = json.dumps(parsed, indent=2)
                        worksheet.cell(row=row, column=col_idx).value = formatted
                        worksheet.cell(row=row, column=col_idx).alignment = Alignment(
                            wrap_text=True,
                            horizontal="center",
                            vertical="center",
                        )
                        lines = formatted.count("\n") + 1
                        current_height = worksheet.row_dimensions[row].height or 0
                        worksheet.row_dimensions[row].height = max(15 * lines, current_height, min_row_height)
                    except Exception:
                        pass
    return workbook


def post_process_workbook_file(file_path: str):
    """Load an existing workbook, post-process it, and resave in place."""
    workbook = load_workbook(file_path)
    post_process_workbook(workbook)
    workbook.save(file_path)
    return workbook
