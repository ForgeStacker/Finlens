"""Primary workbook generation workflow for each AWS profile."""
from __future__ import annotations

import os
import time

import pandas as pd
import boto3
from openpyxl import Workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.styles import Alignment, Border, Font, Side
from openpyxl.utils import get_column_letter

from aws_utils import get_session, sanitize_filename, safe_call
from collectors import COLLECTOR_FUNCTIONS
from config import EXCEL_OUTPUT_DIR, PROFILE_SERVICES, get_account_display_name
import config as cfg


def _config_dir() -> str:
    """Return the absolute path to the project-level config/ directory."""
    # runner/profile_runner.py  ->  runner/  ->  backend/  ->  project root
    return os.path.join(os.path.dirname(os.path.dirname(os.path.dirname(__file__))), "config")


def _read_regions_file() -> list[str]:
    """Read regions from config/regions.yaml (include minus exclude lists)."""
    path = os.path.join(_config_dir(), "regions.yaml")
    try:
        import yaml  # type: ignore
        with open(path, "r") as f:
            data = yaml.safe_load(f) or {}
        included: list[str] = [r for r in (data.get("include") or []) if r]
        excluded: set[str] = set(data.get("exclude") or [])
        regions = [r for r in included if r not in excluded]
        if regions:
            return regions
    except Exception as exc:
        print(f"[WARN] Could not parse regions.yaml: {exc}")
    # Fallback
    try:
        return [cfg.REGION] if cfg.REGION else ["ap-south-1"]
    except Exception:
        return ["ap-south-1"]


def _read_services_file() -> list[str]:
    """Read services from config/services.yaml.

    The YAML uses short lowercase names (e.g. 'ec2', 's3').  We resolve them
    against the keys in COLLECTOR_FUNCTIONS (which may be mixed-case) using a
    case-insensitive match so the caller always gets exact COLLECTOR_FUNCTIONS
    keys back.
    """
    path = os.path.join(_config_dir(), "services.yaml")
    # Build lowercase -> canonical name map once
    lower_map: dict[str, str] = {k.lower(): k for k in COLLECTOR_FUNCTIONS}
    try:
        import yaml  # type: ignore
        with open(path, "r") as f:
            data = yaml.safe_load(f) or {}
        mode: str = data.get("mode", "include")
        listed: list[str] = [str(s) for s in (data.get("list") or []) if s]
        if mode == "include":
            # Return only services that appear in the list (case-insensitive)
            result: list[str] = []
            for name in listed:
                canonical = lower_map.get(name.lower())
                if canonical and canonical not in result:
                    result.append(canonical)
            if result:
                return result
        else:  # exclude
            excluded_lower: set[str] = {s.lower() for s in listed}
            return [k for k in COLLECTOR_FUNCTIONS if k.lower() not in excluded_lower]
    except Exception as exc:
        print(f"[WARN] Could not parse services.yaml: {exc}")
    return list(PROFILE_SERVICES)

from .post_process import post_process_workbook_file
from .utils import format_cell_value, is_cluster_column


# ---------------------------------------------------------------------------
# Per-region output writer
# ---------------------------------------------------------------------------

def _save_region_output(safe_account: str, region: str, region_frames: dict[str, pd.DataFrame]) -> None:
    """Save all service data for one account+region.

    Output layout::

        Data/
          {account}/
            {region}/
              {Service}.csv          <- one CSV per service, normal header row
              {account}_{region}.xlsx <- one workbook, one sheet per service
    """
    import config as _cfg
    region_dir = os.path.join(_cfg.OUTPUT_BASE_DIR, safe_account, sanitize_filename(region))
    os.makedirs(region_dir, exist_ok=True)

    non_empty = {svc: df for svc, df in region_frames.items() if df is not None and not df.empty}

    # ---- CSV: one file per service ----------------------------------------
    for service, df in non_empty.items():
        safe_svc = sanitize_filename(service.replace(" ", "_"))
        csv_path = os.path.join(region_dir, f"{safe_svc}.csv")
        try:
            formatted = df.copy()
            for col in formatted.columns:
                formatted[col] = formatted[col].apply(format_cell_value)
            formatted.to_csv(csv_path, index=False)
            print(f"  [CSV] {csv_path}")
        except Exception as exc:
            print(f"  [CSV] Failed for {service}: {exc}")

    # ---- Excel: one workbook per region, one sheet per service -------------
    workbook = Workbook()
    try:
        workbook.remove(workbook.active)
    except Exception:
        pass

    bold_font = Font(bold=True)
    thin_border = Border(
        left=Side(style="thin", color="000000"),
        right=Side(style="thin", color="000000"),
        top=Side(style="thin", color="000000"),
        bottom=Side(style="thin", color="000000"),
    )

    service_sequence = [svc for svc in PROFILE_SERVICES if svc in non_empty]
    service_sequence += [svc for svc in non_empty if svc not in service_sequence]

    for service in service_sequence:
        df = non_empty.get(service)
        sheet = workbook.create_sheet(service[:31])
        if df is None or df.empty:
            sheet.cell(row=1, column=1, value=f"No data for {service}")
            sheet.row_dimensions[1].height = 24
            continue

        formatted = df.copy()
        for col in formatted.columns:
            formatted[col] = formatted[col].apply(format_cell_value)

        per_sheet_cluster_cols: set[str] = set()
        for col_idx, col_name in enumerate(formatted.columns, 1):
            cell = sheet.cell(row=1, column=col_idx, value=col_name)
            cell.font = bold_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            if is_cluster_column(col_name):
                try:
                    per_sheet_cluster_cols.add(get_column_letter(col_idx))
                except Exception:
                    pass
        sheet.row_dimensions[1].height = 28

        for row_idx, (_, row_vals) in enumerate(formatted.iterrows(), start=2):
            for col_idx, val in enumerate(row_vals, start=1):
                cell = sheet.cell(row=row_idx, column=col_idx, value=val)
                try:
                    col_letter = get_column_letter(col_idx)
                    if col_letter in per_sheet_cluster_cols:
                        cell.alignment = Alignment(wrap_text=False, horizontal="left", vertical="center")
                    else:
                        cell.alignment = Alignment(wrap_text=True, horizontal="center", vertical="center")
                except Exception:
                    cell.alignment = Alignment(wrap_text=True, horizontal="center", vertical="center")
            sheet.row_dimensions[row_idx].height = 24

        last_row = sheet.max_row
        last_col = len(formatted.columns)
        for r in range(1, last_row + 1):
            for c in range(1, last_col + 1):
                sheet.cell(row=r, column=c).border = thin_border

        for column in sheet.columns:
            col_letter = None
            for cell in column:
                if not isinstance(cell, MergedCell):
                    col_letter = cell.column_letter
                    break
            if not col_letter:
                continue
            max_length = 0
            for cell in column:
                try:
                    if cell.value is not None:
                        for line in str(cell.value).split("\n"):
                            if len(line) > max_length:
                                max_length = len(line)
                        if col_letter in per_sheet_cluster_cols:
                            cell.alignment = Alignment(wrap_text=False, horizontal="left", vertical="center")
                        else:
                            cell.alignment = Alignment(wrap_text=True, horizontal="center", vertical="center")
                except Exception:
                    pass
            width = max(12, min(max_length + 4, 120))
            if col_letter in per_sheet_cluster_cols:
                width = max(width, 40)
            sheet.column_dimensions[col_letter].width = width

        for row_idx in range(1, sheet.max_row + 1):
            max_lines = 1
            for col_idx in range(1, sheet.max_column + 1):
                cell = sheet.cell(row=row_idx, column=col_idx)
                if cell.value:
                    lines = str(cell.value).count("\n") + 1
                    if lines > max_lines:
                        max_lines = lines
                try:
                    col_letter = get_column_letter(col_idx)
                    if col_letter in per_sheet_cluster_cols:
                        cell.alignment = Alignment(wrap_text=False, horizontal="left", vertical="center")
                    else:
                        cell.alignment = Alignment(wrap_text=True, horizontal="center", vertical="center")
                except Exception:
                    cell.alignment = Alignment(wrap_text=True, horizontal="center", vertical="center")
            sheet.row_dimensions[row_idx].height = max(15 * max_lines, 18)

    if not workbook.sheetnames:
        ws = workbook.create_sheet("Empty")
        ws.cell(row=1, column=1, value=f"No data collected for {safe_account} / {region}")

    xlsx_name = f"{safe_account}_{sanitize_filename(region)}.xlsx"
    xlsx_path = os.path.join(region_dir, xlsx_name)
    try:
        workbook.save(xlsx_path)
        post_process_workbook_file(xlsx_path)
        print(f"  [Excel] {xlsx_path}")
    except Exception as exc:
        print(f"  [Excel] Failed: {exc}")


# ---------------------------------------------------------------------------
# Main entry point per profile
# ---------------------------------------------------------------------------

def run_for_profile(profile: str) -> None:
    """Collect all services for *profile* across all configured regions.

    Output is written to::

        Data/{account_name}/{region}/{Service}.csv
        Data/{account_name}/{region}/{account}_{region}.xlsx
    """
    print(f"\n=== Processing profile: {profile} ===")
    session = get_session(profile)
    if not session:
        return

    account_name = get_account_display_name(profile, session)
    safe_account = sanitize_filename(account_name.replace(" ", "_"))[:80]

    target_regions = _read_regions_file()
    selected_services = _read_services_file()
    selected_services = [s for s in selected_services if s in COLLECTOR_FUNCTIONS]
    if not selected_services:
        print("services.txt contains no valid collector names. Exiting profile run.")
        return

    for region in target_regions:
        print(f"\n--- Region: {region} ---")
        try:
            cfg.REGION = region
        except Exception:
            pass

        region_frames: dict[str, pd.DataFrame] = {}
        for service in selected_services:
            collector = COLLECTOR_FUNCTIONS.get(service)
            if not collector:
                print(f"  Skipping unknown service: {service}")
                continue
            print(f"  Collecting {service}...")
            dataframe = safe_call(lambda: collector(session, {}), pd.DataFrame())
            if dataframe is None:
                dataframe = pd.DataFrame()
            print(f"  [{service}] shape: {dataframe.shape}")
            region_frames[service] = dataframe
            time.sleep(0.2)

        _save_region_output(safe_account, region, region_frames)
        time.sleep(0.5)

    print(f"\n=== Done: {account_name} ===")


# ---------------------------------------------------------------------------
# Backward-compat shim: old callers pass (profile, workbook) - still works
# ---------------------------------------------------------------------------

def run_for_profile_compat(profile: str, workbook=None) -> None:  # noqa: ARG001
    run_for_profile(profile)

